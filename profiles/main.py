#!/usr/bin/env python3

import getpass
import ipaddress
import intersightpdt
import openpyxl
import pathlib
import re, os, sys, time
import requests
import urllib3
from openpyxl import load_workbook,workbook
from openpyxl.styles import Alignment, colors, Border, Font, NamedStyle, PatternFill, Protection, Side 
from openpyxl.utils.dataframe import dataframe_to_rows
from ordered_set import OrderedSet
from pathlib import Path
from xlutils.copy import copy

func_regex = re.compile('(tfa_|tfr_)')

# Log levels 0 = None, 1 = Class only, 2 = Line
log_level = 2

excel_workbook = None
org_moid = None
api_key = None
home = Path.home()

def stdout_log(sheet, line):
    if log_level == 0:
        return
    elif ((log_level == (1) or log_level == (2)) and
            (sheet) and (line is None)):
        print('*' * 80)
        print('Starting work on {} section'.format(sheet))
        print('*' * 80)
    elif log_level == (2) and (sheet) and (line is not None):
        print('Deploying line {} from section {}...'.format(line, sheet))
    else:
        return


def read_in(excel_workbook):
    try:
        wb = load_workbook(excel_workbook)
        print("Workbook Loaded.")
    except Exception as e:
        print("Something went wrong while opening the workbook - ABORT!")
        sys.exit(e)
    return wb


def findKeys(ws):
    func_list = OrderedSet()
    for i in ws.rows:
        if any(i):
            if re.search(func_regex, str(i[0].value)):
                func_list.add(str(i[0].value))
    return func_list

def countKeys(ws, func):
    count = 0
    for i in ws.rows:
        if any(i):
            if str(i[0].value) == func:
                count += 1
    return count


def findVars(ws, func, count):
    var_list = []
    var_dict = {}
    for i in ws.rows:
        vcount = 1
        if any(i):
            if str(i[0].value) == func:
                try:
                    for x in range(2, 17):
                        if str(i[i - 1, x].value):
                            var_list.append(str(i[i - vcount, x].value))
                            print(var_list)
                except Exception as e:
                    e = e
                    pass
                break
        else:
            vcount += 1
    while count > 0:
        var_dict[count] = {}
        var_count = 0
        for z in var_list:
            var_dict[count][z] = str(i[i + count - 1, 4 + var_count].value)
            var_count += 1
        var_dict[count]['row'] = i + count - 1
        count -= 1
    return var_dict

def wb_update(wr_ws, status, i):
    # build green and red style sheets for excel
    bd1 = Side(style="thick", color="8EA9DB")
    bd2 = Side(style="medium", color="8EA9DB")
    wsh1 = NamedStyle(name="wsh1")
    wsh1.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
    wsh1.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
    wsh1.font = Font(bold=True, size=15, color="FFFFFF")
    wsh2 = NamedStyle(name="wsh2")
    wsh2.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
    wsh2.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    wsh2.fill = PatternFill("solid", fgColor="305496")
    wsh2.font = Font(bold=True, size=15, color="FFFFFF")
    green_st = NamedStyle(name="ws_odd")
    green_st.alignment = Alignment(horizontal="center", vertical="center")
    green_st.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    green_st.fill = PatternFill("solid", fgColor="D9E1F2")
    green_st.font = Font(bold=False, size=12, color="44546A")
    red_st = NamedStyle(name="ws_even")
    red_st.alignment = Alignment(horizontal="center", vertical="center")
    red_st.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    red_st.font = Font(bold=False, size=12, color="44546A")
    yellow_st = NamedStyle(name="ws_even")
    yellow_st.alignment = Alignment(horizontal="center", vertical="center")
    yellow_st.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    yellow_st.font = Font(bold=False, size=12, color="44546A")
    # green_st = xlwt.easyxf('pattern: pattern solid;')
    # green_st.pattern.pattern_fore_colour = 3
    # red_st = xlwt.easyxf('pattern: pattern solid;')
    # red_st.pattern.pattern_fore_colour = 2
    # yellow_st = xlwt.easyxf('pattern: pattern solid;')
    # yellow_st.pattern.pattern_fore_colour = 5
    # if stanzas to catch the status code from the request
    # and then input the appropriate information in the workbook
    # this then writes the changes to the doc
    if status == 200:
        wr_ws.write(i, 1, 'Success (200)', green_st)
    if status == 400:
        print("Error 400 - Bad Request - ABORT!")
        print("Probably have a bad URL or payload")
        wr_ws.write(i, 1, 'Bad Request (400)', red_st)
        pass
    if status == 401:
        print("Error 401 - Unauthorized - ABORT!")
        print("Probably have incorrect credentials")
        wr_ws.write(i, 1, 'Unauthorized (401)', red_st)
        pass
    if status == 403:
        print("Error 403 - Forbidden - ABORT!")
        print("Server refuses to handle your request")
        wr_ws.write(i, 1, 'Forbidden (403)', red_st)
        pass
    if status == 404:
        print("Error 404 - Not Found - ABORT!")
        print("Seems like you're trying to POST to a page that doesn't"
              " exist.")
        wr_ws.write(i, 1, 'Not Found (400)', red_st)
        pass
    if status == 666:
        print("Error - Something failed!")
        print("The POST failed, see stdout for the exception.")
        wr_ws.write(i, 1, 'Unkown Failure', yellow_st)
        pass
    if status == 667:
        print("Error - Invalid Input!")
        print("Invalid integer or other input.")
        wr_ws.write(i, 1, 'Unkown Failure', yellow_st)
        pass

def server_profile(api_key, priv_key, org_moid, wb):
    ws = wb['Service Profile']
    func_list = findKeys(ws)
    print(func_list)
    srvprf = intersightpdt.config_server(api_key, priv_key, org_moid, wb)
    # stdout_log(wr_ws.name, None)
    for func in func_list:
        count = countKeys(ws, func)
        var_dict = findVars(ws, func, count)
        for pos in var_dict:
            # row_num = var_dict[pos]['row']
            # del var_dict[pos]['row']
            for x in list(var_dict[pos].keys()):
                if var_dict[pos][x] == '':
                    del var_dict[pos][x]
            #stdout_log(wr_ws.name, row_num)
            status = eval("%s.%s(**var_dict[pos])" % (srvprf, func))
            print(status)
            #wb_update(wr_ws, status, row_num)
            time.sleep(.025)

def main():
    # Disable urllib3 warnings
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    # Ask user for required Information: INTERSIGHT_DEPLOY_FILE, INTERSIGHT_ORG_MOID, INTERSIGHT_PRIVATE_KEY 
    if sys.argv[1:]:
        if os.path.isfile(sys.argv[1]):
            excel_workbook = sys.argv[1]
        else:
            print('\nWorkbook not Found.  Please enter a valid /path/filename for the source you will be using.')
            while True:
                print('Please enter a valid /path/filename for the source you will be using.')
                excel_workbook = input('/Path/Filename: ')
                if os.path.isfile(excel_workbook):
                    print(f'\n-----------------------------------------------------------------------------\n')
                    print(f'   {excel_workbook} exists.  Will Now Check for API Variables...')
                    print(f'\n-----------------------------------------------------------------------------\n')
                    break
                else:
                    print('\nWorkbook not Found.  Please enter a valid /path/filename for the source you will be using.')
    else:
        while True:
            print('Please enter a valid /path/filename for the source you will be using.')
            excel_workbook = input('/Path/Filename: ')
            if os.path.isfile(excel_workbook):
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'   {excel_workbook} exists.  Will Now Check for API Variables...')
                print(f'\n-----------------------------------------------------------------------------\n')
                break
            else:
                print('\nWorkbook not Found.  Please enter a valid /path/filename for the source you will be using.')

    # Collect API Variables for Automation
    api_regex = re.compile(r'^API_KEY . \"(.{74})\"$')
    api_key = ''
    api_file_count = 0
    if sys.argv[2:]:
        if os.path.isfile(sys.argv[2]):
            api_vars_file = sys.argv[2]
            api_file_count +=1
        else:
            print('\nAPI Variable File not Found.  Please enter a valid /path/filename for the API Variables.')
            while True:
                print('Please enter a valid /path/filename for the API Variables.')
                api_vars_file = input('/Path/Filename: ')
                if os.path.isfile(api_vars_file):
                    print(f'\n-----------------------------------------------------------------------------\n')
                    print(f'   {api_vars_file} exists.  Need to Check for the Private Key...')
                    print(f'\n-----------------------------------------------------------------------------\n')
                    api_file_count +=1
                    break
                else:
                    print('\nAPI Variable File not Found.  Please enter a valid /path/filename for the API Variables.')
    else:
        while True:
            print('Please enter a valid /path/filename for the API Variables.')
            api_vars_file = input('/Path/Filename: ')
            if os.path.isfile(api_vars_file):
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'   {api_vars_file} exists.  Need to Check for the Private Key...')
                print(f'\n-----------------------------------------------------------------------------\n')
                api_file_count +=1
                break
            else:
                print('\nAPI Variable File not Found.  Please enter a valid /path/filename for the API Variables.')

    if not api_file_count == 0:
        api_line_count = 0
        read_file = open(api_vars_file, 'r')
        for line in read_file:
            if  re.search(api_regex, line):
                api_key = re.search(api_regex, line).group(1)
                api_line_count += 1

        if not api_line_count == 1:
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'   Did not find The proper Varialbles in the API Variables File.')
                print(f'   Expected the following:\n')
                print(f'        API_KEY = "some_string"')
                print(f'\n   Existing Script...')
                print(f'\n-----------------------------------------------------------------------------\n')
                exit()

    # Obtain the Private Key File
    priv_key = ''
    if sys.argv[3:]:
        if os.path.isfile(sys.argv[3]):
            priv_key = sys.argv[3]
        else:
            print('\nPrivate Key File not Found.  Please enter a valid /path/filename for the Private Key.')
            while True:
                print('Please enter a valid /path/filename for the Private Key.')
                priv_key = input('/Path/Filename: ')
                if os.path.isfile(priv_key):
                    print(f'\n-----------------------------------------------------------------------------\n')
                    print(f'   {priv_key} exists.  Beginning Script Execution...')
                    print(f'\n-----------------------------------------------------------------------------\n')
                    break
                else:
                    print('\nPrivate Key File not Found.  Please enter a valid /path/filename for the Private Key.')
    else:
        while True:
            print('Please enter a valid /path/filename for the Private Key.')
            priv_key = input('/Path/Filename: ')
            if os.path.isfile(priv_key):
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'   {priv_key} exists.  Beginning Script Execution...')
                print(f'\n-----------------------------------------------------------------------------\n')
                break
            else:
                print('\nPrivate Key File not Found.  Please enter a valid /path/filename for the Private Key.')

    # Load Workbook
    wb = read_in(excel_workbook)
    # Load a Copy Workbook
    #wr_wb = copy(wb)
    server_profile(api_key, priv_key, org_moid, wb)



if __name__ == '__main__':
    main()